#! python 3
# _*_ coding: utf_8  _*_

'''
During the course, note the performance for the students.
select the "课程"，select "加减分的项",input"学号","分值"
it will write the "分值"
Record can not be written. prompt is not good enough.   2017-2-11-10:50
Student's number must be 2 digits. When it is smaller than 10, It should be 0X.  2017-2-11-17:23
display the 总分。 design for write the mark of the homework.    2017-2-20
'''

import openpyxl
import os
import re
import logging

# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

# find the Chinese words '学生名单' in filename
ChineseReg = re.compile(r'学生名单')
# find the class
classReg = re.compile(r'\d{7}')
# find the course
courseReg = re.compile(r'-(\w{3,11})-')

performanceTag = [
    ['学号', ],
    ['姓名', ],
    ['总分',],['初始分',],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['提出问题',],
    ['回答问题',],
    ['作业1',],['作业2',],['作业3',],['作业4',],['作业5',],['作业6',],['作业7',],['作业8',],
    ]
labTag = [
    ['学号', ],
    ['姓名', ],
    ['总分',],['初始分',],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],
    ['报告1',],['报告2',],['报告3',],['报告4',],    
    ]
designTag = [
    ['学号', ],
    ['姓名', ],
    ['总分',],['初始分',],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['设计1',],['设计2',],['设计3',],['设计4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],
    ['报告1',],['报告2',],    
    ]
practiceTag = [
    ['学号', ],
    ['姓名', ],
    ['总分',],['初始分',],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],
    ['报告1',],['报告2',],    
    ]

# 提示输入信息
filelist = []
fulllist = []
k = 0
dirname = 'd:\\_PythonWorks\\excelOperate\\pscj161702'
for file in os.listdir(dirname):
    fullname = dirname + '\\' + file
    if not ChineseReg.search(file):    # ChineseReg = re.compile(r'学生名单')
        filelist.append(file)
        fulllist.append(fullname)     
        print(k,file)
        k += 1
courseNum = int(input('please input a number for 课程: '))

tagdict = {'performance':performanceTag, 'lab':labTag, 'design':designTag, 'practice':practiceTag}
coursetype = courseReg.search(filelist[courseNum]).group(1)
logging.info(coursetype)
print(coursetype)
item = {}
k = 2
for val in tagdict[coursetype]:
    print(str(k) + ' ' + str(val) + ' ')
    item[k] = str(val)
    k += 1            
itemNum = int(input('please input a numbers for select item: '))
        
wb = openpyxl.load_workbook(fulllist[courseNum])
sheet = wb.get_active_sheet()
finish = 0
while not finish:
    stuNum = input("\n please input two last digitals of select student's number: 05 ")
    mark = input('\n please input the mark: ')
    
    for row in range(3,sheet.max_row + 1):
        logging.debug(str(sheet['b'+str(row)].value)[-2:])           #  学号在B列
        if str(sheet['b'+str(row)].value)[-2:] == stuNum:                   #  学号在B列
            # Write
            logging.critical(sheet.cell(row = row,column = itemNum).value)   # 写之前，cell的值
            sheet.cell(row = row,column = itemNum).value += int(mark)        # 加上要加减的分数
            sheet.cell(row = row,column = 4).value += int(mark)              # 总分也加上该分数
            logging.critical(sheet.cell(row = row,column = itemNum).value)   # 写之后，cell的值
            print(stuNum,' 总分是：', sheet.cell(row = row,column = 4).value)
            break
    finish = input("input any letter to finish. display all students' mark, then end.")
marks = []
for row in range(3,sheet.max_row + 1):
    marks.append((sheet.cell(row = row,column = 4).value, sheet['b'+str(row)].value))
wb.save(fulllist[courseNum])
marks.sort(reverse=True)
print('前8名为：', marks[:8])
logging.critical('---------------')
print('后8名为：', marks[-8:])

logging.critical('-------End--------')






