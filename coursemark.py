#! python 3
# _*_ coding: utf_8  _*_

'''
During the course, note the performance for the students.
select the "课程"，select "加减分的项",input"学号","分值"
it will write the "分值"
Record can not be written. prompt is not good enough.   2017-2-11-10:50
Student's number must be 2 digits. When it is smaller than 10, It should be 0X.  2017-2-11-17:23
tidy the value. add display the 总分。
'''

import openpyxl
import os
import re
import logging

# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

# find the Chinese words '学生名单' in filename
Chinese_reg = re.compile(r'学生名单')
# find the class
class_reg = re.compile(r'\d{7}')
# find the course
course_reg = re.compile(r'-(\w{3,11})-')

performance_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['提出问题',],
    ['回答问题',],
    ]
lab_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],   
    ]
design_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['设计1',],['设计2',],['设计3',],['设计4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],  
    ]
practice_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],   
    ]

# 提示输入信息
filelist = []
fulllist = []
k = 0
DIRNAME = 'd:\\_PythonWorks\\excelOperate\\pscj161702'
for file in os.listdir(DIRNAME):
    fullname = DIRNAME + '\\' + file
    if not Chinese_reg.search(file):    # Chinese_reg = re.compile(r'学生名单')
        filelist.append(file)
        fulllist.append(fullname)     
        print(k,file)
        k += 1      
tagdict = {'performance':performance_tag, 'lab':lab_tag, 'design':design_tag, 'practice':practice_tag}
coursenum = int(input('\n please input a number for 课程: '))    
coursetype = course_reg.search(filelist[coursenum]).group(1)
finish = 0
while finish != 'y':  
    print(coursetype)
    item = {}
    k = 6
    for val in tagdict[coursetype]:
        print(str(k) + ' ' + str(val) + ' ')
        item[k] = str(val)
        k += 1

    itemnum = int(input('\n please input a number for select item: '))
    studnum = input("\n please input three last digitals of select student's number: 205 ")
    mark = input('\n please input the mark: ')
    
    wb = openpyxl.load_workbook(fulllist[coursenum])
    sheet = wb.get_active_sheet()
    
    for row in range(3,sheet.max_row + 1):
        logging.debug(str(sheet['b'+str(row)].value)[-3:])           #  学号在B列
        if str(sheet['b'+str(row)].value)[-3:] == studnum:                   #  学号在B列
            # Write
            logging.critical(sheet.cell(row = row,column = itemnum).value)   # 写之前，cell的值
            sheet.cell(row = row,column = itemnum).value += int(mark)        # 加上要加减的分数
            sheet.cell(row = row,column = 4).value += int(mark)              # 总分也加上该分数
            logging.critical(sheet.cell(row = row,column = itemnum).value)   # 写之后，cell的值
            print(studnum,' 总分是：', sheet.cell(row = row,column = 4).value)
            break
        
    marks = []
    for row in range(3,sheet.max_row + 1):
        marks.append((sheet.cell(row = row,column = 4).value, sheet['b'+str(row)].value, sheet['c'+str(row)].value))
        
    wb.save(fulllist[coursenum])
    
    marks.sort(reverse=True)
    print('前8名为：')
    for k in marks[:8]:
        print(k)
    logging.critical('---------------')
    finish = input('你要退出吗？(y/n):').lower()

##print('后5名为：')
##for k in marks[-5:]:
##    print(k)

logging.critical('-------End--------')






